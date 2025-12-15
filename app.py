"""
Sistema de Gestión Inteligente de Laboratorios (GIL)
Centro Minero de Sogamoso - SENA
Archivo principal de la aplicación web
"""

import sys
import os

# Configurar codificación UTF-8 para Windows
if sys.platform == 'win32':
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

from flask import Flask, render_template, redirect, url_for, request, jsonify, session, flash, send_file
from flask_cors import CORS
from flask_mail import Mail, Message
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Agregar rutas al path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Importar configuración centralizada
from config.config import Config, get_config
from config.api_config import APIConfig

# Importar módulos del backend
from backend.utils.database import DatabaseManager
from backend.utils.auth import AuthManager

# Obtener configuración según entorno
config = get_config()

# Configuración de la aplicación
app = Flask(__name__, 
            template_folder='frontend/templates',
            static_folder='frontend/static')

# Aplicar configuración desde Config centralizada
app.config['SECRET_KEY'] = Config.SECRET_KEY
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = Config.PERMANENT_SESSION_LIFETIME
app.config['JSON_AS_ASCII'] = False
app.config['JSON_SORT_KEYS'] = False

# Deshabilitar caché en desarrollo
if Config.FLASK_DEBUG:
    app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
    app.config['TEMPLATES_AUTO_RELOAD'] = True

# Habilitar CORS
CORS(app, origins=APIConfig.API_CORS_ORIGINS)

# Configurar Flask-Mail para envío de emails
app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.getenv('MAIL_USE_TLS', 'True') == 'True'
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_DEFAULT_SENDER', 'noreply@centrominero.edu.co')

mail = Mail(app)

# Inicializar base de datos con Config
db_manager = DatabaseManager()

# Intentar registrar blueprints de la API
try:
    from backend.api.blueprints import registrar_blueprints
    registrar_blueprints(app)
except ImportError as e:
    print(f"⚠️  No se pudieron cargar los blueprints: {e}")

# Ruta para servir archivos de uploads (imágenes de equipos)
@app.route('/uploads/<path:filename>')
def serve_uploads(filename):
    """Servir archivos de la carpeta uploads"""
    from flask import send_from_directory
    uploads_dir = os.path.join(os.path.dirname(__file__), 'uploads')
    return send_from_directory(uploads_dir, filename)

# ========================================
# MIDDLEWARE - Deshabilitar caché en desarrollo
# ========================================

@app.after_request
def add_no_cache_headers(response):
    """Agregar headers para evitar caché en desarrollo"""
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# ========================================
# FILTROS PERSONALIZADOS JINJA2
# ========================================

from datetime import datetime

@app.template_filter('strptime')
def strptime_filter(date_string, format='%d/%m/%Y'):
    """Convierte string a datetime"""
    try:
        return datetime.strptime(date_string, format)
    except:
        return None

# Agregar función moment() como global en Jinja2
def moment():
    """Retorna la fecha actual como objeto datetime"""
    return datetime.now()

app.jinja_env.globals.update(moment=moment)

# ========================================
# FUNCIONES AUXILIARES
# ========================================

def get_user_data():
    """Obtiene datos del usuario actual de la sesión incluyendo permisos"""
    permisos = session.get('user_permisos', {})
    return {
        'user_id': session.get('user_id'),
        'user_name': session.get('user_name', 'Usuario'),
        'user_type': session.get('user_type', 'usuario'),
        'user_level': session.get('user_level', 1),
        'user_rol': session.get('user_rol'),  # id_rol: 1=Admin, 2=Instructor, 3=Técnico, 4=Aprendiz, 5=Coordinador
        'user_documento': session.get('user_documento', ''),
        'permisos': permisos  # Permisos del rol desde la BD
    }

def tiene_permiso(modulo, incluir_ver=True):
    """
    Verifica si el usuario actual tiene permiso para un módulo.
    Si incluir_ver=True, también verifica permisos de solo lectura (modulo_ver)
    """
    permisos = session.get('user_permisos', {})
    if permisos.get('all'):
        return True
    if permisos.get(modulo):
        return True
    if incluir_ver and permisos.get(f'{modulo}_ver'):
        return True
    return False

def puede_editar(modulo):
    """Verifica si el usuario puede editar (no solo ver) un módulo"""
    permisos = session.get('user_permisos', {})
    if permisos.get('all'):
        return True
    return permisos.get(modulo, False)

def obtener_estadisticas_sistema():
    """
    Obtiene estadísticas del sistema para dashboard y API.
    Queries adaptadas a schema.sql
    
    Returns:
        dict: Diccionario con todas las estadísticas del sistema
    """
    stats = {
        'equipos_estado': {
            'disponible': 0,
            'prestado': 0,
            'mantenimiento': 0,
            'reparacion': 0,
            'dado_baja': 0
        },
        'inventario_critico': 0,
        'reservas_activas': 0,
        'usuarios_activos_hoy': 0,
        'comandos_hoy': 0
    }
    
    try:
        # Contar equipos por estado (columna 'estado' según schema.sql)
        query_equipos = """
            SELECT estado, COUNT(*) as total 
            FROM equipos 
            GROUP BY estado
        """
        resultados = db_manager.ejecutar_query(query_equipos)
        for row in resultados:
            estado = row.get('estado', '')
            total = row.get('total', 0)
            if estado in stats['equipos_estado']:
                stats['equipos_estado'][estado] = total
        
        # Contar inventario crítico - Nota: tabla puede no existir
        try:
            query_inventario = """
                SELECT COUNT(*) as total 
                FROM inventario 
                WHERE cantidad_actual <= cantidad_minima
            """
            resultado = db_manager.obtener_uno(query_inventario)
            if resultado:
                stats['inventario_critico'] = resultado.get('total', 0)
        except:
            pass
        
        # Contar préstamos activos (columna 'estado' según schema.sql)
        query_prestamos = """
            SELECT COUNT(*) as total 
            FROM prestamos 
            WHERE estado = 'activo'
        """
        resultado = db_manager.obtener_uno(query_prestamos)
        if resultado:
            stats['reservas_activas'] = resultado.get('total', 0)
        
        # Contar usuarios activos (columna 'estado' según schema.sql)
        query_usuarios = """
            SELECT COUNT(DISTINCT id) as total 
            FROM usuarios 
            WHERE estado = 'activo'
        """
        resultado = db_manager.obtener_uno(query_usuarios)
        if resultado:
            stats['usuarios_activos_hoy'] = resultado.get('total', 0)
            
    except Exception as e:
        print(f"Error obteniendo estadísticas: {e}")
    
    return stats

def enviar_email_reset(email, reset_url, nombre):
    """Enviar email con enlace de restablecimiento de contraseña"""
    try:
        msg = Message(
            'Restablecer Contraseña - Centro Minero SENA',
            recipients=[email]
        )
        msg.html = f'''
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: #667eea; color: white; padding: 30px; text-align: center; border-radius: 10px 10px 0 0; }}
                .content {{ padding: 30px; background: #f9f9f9; }}
                .button {{ background: #667eea; color: white; padding: 15px 40px; text-decoration: none; 
                          border-radius: 5px; display: inline-block; font-weight: bold; }}
                .footer {{ text-align: center; padding: 20px; color: #666; font-size: 12px; }}
                .warning {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1 style="margin: 0;">🔐 Restablecer Contraseña</h1>
                </div>
                <div class="content">
                    <p>Hola <strong>{nombre}</strong>,</p>
                    <p>Recibimos una solicitud para restablecer tu contraseña en el Sistema de Gestión Integral de Laboratorios (GIL).</p>
                    <p style="text-align: center; margin: 30px 0;">
                        <a href="{reset_url}" class="button">Restablecer Mi Contraseña</a>
                    </p>
                    <div class="warning">
                        <strong>⏰ Este enlace expirará en 1 hora.</strong>
                    </div>
                    <p>Si no solicitaste este cambio, ignora este correo y tu contraseña permanecerá sin cambios.</p>
                    <p style="color: #666; font-size: 14px;">Por seguridad, nunca compartas este enlace con nadie.</p>
                </div>
                <div class="footer">
                    <p><strong>Centro Minero de Sogamoso - SENA</strong></p>
                    <p>Sistema de Gestión Integral de Laboratorios (GIL)</p>
                    <p style="color: #999;">Este es un correo automático, por favor no responder.</p>
                </div>
            </div>
        </body>
        </html>
        '''
        mail.send(msg)
        return True
    except Exception as e:
        print(f"❌ Error enviando email: {e}")
        import traceback
        traceback.print_exc()
        return False

# ========================================
# RUTAS PRINCIPALES
# ========================================

@app.route('/')
def index():
    """Página de inicio - redirige al login"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login con validación de contraseña"""
    if request.method == 'POST':
        try:
            user_id = request.form.get('user_id', '').strip()
            password = request.form.get('password', '')
            
            if not user_id:
                flash('Por favor ingrese su número de documento', 'error')
                return render_template('login.html')
            
            if not password:
                flash('Por favor ingrese su contraseña', 'error')
                return render_template('login.html')
            
            # Buscar usuario en BD con permisos del rol
            query = """
                SELECT u.id, u.documento, u.nombres, u.apellidos, u.email,
                       u.id_rol, u.estado, u.password_hash, r.nombre_rol, r.permisos
                FROM usuarios u
                LEFT JOIN roles r ON u.id_rol = r.id
                WHERE (u.id = %s OR u.documento = %s) AND u.estado = 'activo'
            """
            usuario = db_manager.obtener_uno(query, (user_id, user_id))
            
            if not usuario:
                flash('Usuario no encontrado o inactivo', 'error')
                return render_template('login.html')
            
            # Validar contraseña con bcrypt
            password_hash = usuario.get('password_hash', '')
            if not password_hash:
                flash('Usuario sin contraseña configurada. Contacte al administrador.', 'error')
                return render_template('login.html')
            
            try:
                import bcrypt
                password_valid = bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))
            except Exception as e:
                print(f"Error verificando contraseña: {e}")
                password_valid = False
            
            if not password_valid:
                flash('Contraseña incorrecta', 'error')
                # Registrar intento fallido
                try:
                    log_query = """
                        INSERT INTO logs_sistema (modulo, nivel_log, mensaje, id_usuario, ip_address)
                        VALUES ('auth', 'WARNING', 'Intento de login fallido - contraseña incorrecta', %s, %s)
                    """
                    db_manager.ejecutar_comando(log_query, (usuario['id'], request.remote_addr))
                except:
                    pass
                return render_template('login.html')
            
            # Calcular nivel de acceso según rol
            niveles_rol = {1: 5, 2: 4, 3: 3, 4: 2, 5: 2, 6: 1}  # admin=5, instructor=4, etc.
            nivel = niveles_rol.get(usuario['id_rol'], 1)
            nombre_completo = f"{usuario['nombres']} {usuario['apellidos']}"
            
            # Parsear permisos del rol desde JSON
            import json
            permisos_str = usuario.get('permisos', '{}') or '{}'
            try:
                permisos = json.loads(permisos_str)
            except:
                permisos = {}
            
            # Login exitoso - crear sesión
            session['user_id'] = usuario['id']
            session['user_name'] = nombre_completo
            session['user_type'] = usuario['nombre_rol'] or 'usuario'
            session['user_level'] = nivel
            session['user_rol'] = usuario['id_rol']  # id_rol de la BD: 1=Admin, 2=Instructor, 3=Técnico, 4=Aprendiz, 5=Coordinador
            session['user_permisos'] = permisos  # Permisos del rol desde la BD
            session['user_documento'] = usuario['documento']
            session.permanent = True
            
            # Actualizar último acceso
            try:
                db_manager.ejecutar_comando(
                    "UPDATE usuarios SET ultimo_acceso = NOW() WHERE id = %s",
                    (usuario['id'],)
                )
            except:
                pass
            
            # Registrar login exitoso
            try:
                log_query = """
                    INSERT INTO logs_sistema (modulo, nivel_log, mensaje, id_usuario, ip_address)
                    VALUES ('auth', 'INFO', 'Login exitoso', %s, %s)
                """
                db_manager.ejecutar_comando(log_query, (usuario['id'], request.remote_addr))
            except:
                pass
            
            flash(f'Bienvenido {nombre_completo}!', 'success')
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            print(f"Error en login: {e}")
            flash('Error al iniciar sesión. Intente nuevamente.', 'error')
            return render_template('login.html')
    
    return render_template('login.html')

@app.route('/forgot-password', methods=['POST'])
def forgot_password():
    """Solicitar restablecimiento de contraseña"""
    try:
        import secrets
        from datetime import datetime, timedelta
        
        email = request.json.get('email', '').strip().lower()
        
        if not email:
            return jsonify({'success': False, 'message': 'Email requerido'}), 400
        
        # Buscar usuario por email
        query = "SELECT id, documento, nombres, apellidos, email FROM usuarios WHERE LOWER(email) = %s AND estado = 'activo'"
        usuario = db_manager.obtener_uno(query, (email,))
        
        # Por seguridad, siempre responder lo mismo aunque el email no exista
        if not usuario:
            return jsonify({
                'success': True, 
                'message': 'Si el correo está registrado, recibirá instrucciones en breve.'
            })
        
        # Generar token único y seguro
        token = secrets.token_urlsafe(32)
        expira_en = datetime.now() + timedelta(hours=1)  # Token válido por 1 hora
        
        # Guardar token en BD
        insert_query = """
            INSERT INTO password_reset_tokens (id_usuario, token, email, expira_en, ip_solicitud)
            VALUES (%s, %s, %s, %s, %s)
        """
        db_manager.ejecutar_comando(insert_query, (
            usuario['id'],
            token,
            email,
            expira_en,
            request.remote_addr
        ))
        
        # Construir URL de restablecimiento
        reset_url = f"{request.url_root}reset-password/{token}"
        
        # Enviar email con el enlace
        email_enviado = enviar_email_reset(
            email, 
            reset_url, 
            f"{usuario['nombres']} {usuario['apellidos']}"
        )
        
        if email_enviado:
            print(f"✅ Email enviado exitosamente a {email}")
        else:
            print(f"⚠️  No se pudo enviar el email, pero el token fue generado")
            print(f"=== TOKEN DE RESTABLECIMIENTO ===")
            print(f"Usuario: {usuario['nombres']} {usuario['apellidos']}")
            print(f"Email: {email}")
            print(f"URL: {reset_url}")
            print(f"Expira: {expira_en}")
            print(f"================================")
        
        # Log en sistema
        log_query = """
            INSERT INTO logs_sistema (modulo, nivel_log, mensaje, id_usuario, ip_address)
            VALUES ('auth', 'INFO', %s, %s, %s)
        """
        db_manager.ejecutar_comando(log_query, (
            f'Solicitud de restablecimiento de contraseña para {email}',
            usuario['id'],
            request.remote_addr
        ))
        
        return jsonify({
            'success': True,
            'message': 'Si el correo está registrado, recibirá instrucciones en breve.',
            'debug_url': reset_url if app.debug else None  # Solo en desarrollo
        })
        
    except Exception as e:
        print(f"Error en forgot_password: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Error al procesar solicitud'}), 500

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    """Restablecer contraseña con token"""
    from datetime import datetime
    
    if request.method == 'GET':
        # Verificar que el token sea válido
        query = """
            SELECT t.id, t.id_usuario, t.email, t.expira_en, t.usado,
                   u.nombres, u.apellidos, u.documento
            FROM password_reset_tokens t
            JOIN usuarios u ON t.id_usuario = u.id
            WHERE t.token = %s AND t.usado = FALSE
        """
        token_data = db_manager.obtener_uno(query, (token,))
        
        if not token_data:
            flash('Token inválido o ya utilizado', 'error')
            return redirect(url_for('login'))
        
        # Verificar si expiró
        if datetime.now() > token_data['expira_en']:
            flash('El enlace ha expirado. Solicite uno nuevo.', 'error')
            return redirect(url_for('login'))
        
        # Mostrar formulario de nueva contraseña
        return render_template('reset_password.html', 
                             token=token,
                             email=token_data['email'],
                             nombre=f"{token_data['nombres']} {token_data['apellidos']}")
    
    # POST - Cambiar contraseña
    try:
        import re
        
        nueva_password = request.form.get('nueva_password', '').strip()
        confirmar_password = request.form.get('confirmar_password', '').strip()
        
        # Validación 1: Campos requeridos
        if not nueva_password or not confirmar_password:
            flash('Todos los campos son requeridos', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Validación 2: Coincidencia de contraseñas
        if nueva_password != confirmar_password:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Validación 3: Longitud mínima (8 caracteres)
        if len(nueva_password) < 8:
            flash('La contraseña debe tener al menos 8 caracteres', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Validación 4: Al menos una letra mayúscula
        if not re.search(r'[A-Z]', nueva_password):
            flash('La contraseña debe contener al menos una letra mayúscula (A-Z)', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Validación 5: Al menos una letra minúscula
        if not re.search(r'[a-z]', nueva_password):
            flash('La contraseña debe contener al menos una letra minúscula (a-z)', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Validación 6: Al menos un número
        if not re.search(r'[0-9]', nueva_password):
            flash('La contraseña debe contener al menos un número (0-9)', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Validación 7: Al menos un carácter especial
        if not re.search(r'[!@#$%^&*(),.?":{}|<>_\-]', nueva_password):
            flash('La contraseña debe contener al menos un carácter especial (!@#$%^&*...)', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Verificar token nuevamente
        query = """
            SELECT t.id, t.id_usuario, t.email, t.expira_en, t.usado
            FROM password_reset_tokens t
            WHERE t.token = %s AND t.usado = FALSE
        """
        token_data = db_manager.obtener_uno(query, (token,))
        
        if not token_data:
            flash('Token inválido o ya utilizado', 'error')
            return redirect(url_for('login'))
        
        if datetime.now() > token_data['expira_en']:
            flash('El enlace ha expirado. Solicite uno nuevo.', 'error')
            return redirect(url_for('login'))
        
        # Hashear nueva contraseña
        import bcrypt
        password_hash = bcrypt.hashpw(nueva_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        
        # Actualizar contraseña
        update_query = "UPDATE usuarios SET password_hash = %s WHERE id = %s"
        db_manager.ejecutar_comando(update_query, (password_hash, token_data['id_usuario']))
        
        # Marcar token como usado
        mark_used_query = "UPDATE password_reset_tokens SET usado = TRUE WHERE id = %s"
        db_manager.ejecutar_comando(mark_used_query, (token_data['id'],))
        
        # Log de seguridad
        log_query = """
            INSERT INTO logs_sistema (modulo, nivel_log, mensaje, id_usuario, ip_address)
            VALUES ('auth', 'INFO', %s, %s, %s)
        """
        db_manager.ejecutar_comando(log_query, (
            f'Contraseña restablecida exitosamente para {token_data["email"]}',
            token_data['id_usuario'],
            request.remote_addr
        ))
        
        flash('Contraseña restablecida exitosamente. Ahora puede iniciar sesión.', 'success')
        return redirect(url_for('login'))
        
    except Exception as e:
        print(f"Error en reset_password: {e}")
        import traceback
        traceback.print_exc()
        flash('Error al restablecer contraseña. Intente nuevamente.', 'error')
        return redirect(url_for('reset_password', token=token))

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Página de registro de nuevos usuarios - Validaciones híbridas"""
    if request.method == 'POST':
        try:
            documento = request.form.get('documento', '').strip()
            nombres = request.form.get('nombres', '').strip()
            apellidos = request.form.get('apellidos', '').strip()
            email = request.form.get('email', '').strip().lower()
            telefono = request.form.get('telefono', '').strip()
            password = request.form.get('password', '')
            password_confirm = request.form.get('password_confirm', '')
            
            # Validar campos obligatorios
            if not all([documento, nombres, apellidos, email, password]):
                flash('Por favor complete todos los campos obligatorios', 'error')
                return render_template('register.html')
            
            # Validar confirmación de contraseña
            if password != password_confirm:
                flash('Las contraseñas no coinciden', 'error')
                return render_template('register.html')
            
            # Validaciones híbridas - Backend
            import re
            
            # Validar documento (6-20 dígitos)
            if not re.match(r'^[0-9]{6,20}$', documento):
                flash('El documento debe tener entre 6 y 20 dígitos numéricos', 'error')
                return render_template('register.html')
            
            # Validar nombres (solo letras y espacios, 2-100 caracteres)
            if not re.match(r'^[a-záéíóúñA-ZÁÉÍÓÚÑ\s]{2,100}$', nombres):
                flash('Los nombres solo pueden contener letras y espacios (2-100 caracteres)', 'error')
                return render_template('register.html')
            
            # Validar apellidos (solo letras y espacios, 2-100 caracteres)
            if not re.match(r'^[a-záéíóúñA-ZÁÉÍÓÚÑ\s]{2,100}$', apellidos):
                flash('Los apellidos solo pueden contener letras y espacios (2-100 caracteres)', 'error')
                return render_template('register.html')
            
            # Validar formato de email
            email_regex = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_regex, email):
                flash('Formato de email inválido', 'error')
                return render_template('register.html')
            
            # Validar teléfono si se proporciona (7-15 dígitos)
            if telefono and not re.match(r'^[0-9]{7,15}$', telefono):
                flash('El teléfono debe tener entre 7 y 15 dígitos', 'error')
                return render_template('register.html')
            
            # Validar contraseña segura (mínimo 8 caracteres, mayúscula, minúscula, número, carácter especial)
            if len(password) < 8:
                flash('La contraseña debe tener al menos 8 caracteres', 'error')
                return render_template('register.html')
            
            if not re.search(r'[A-Z]', password):
                flash('La contraseña debe contener al menos una letra mayúscula', 'error')
                return render_template('register.html')
            
            if not re.search(r'[a-z]', password):
                flash('La contraseña debe contener al menos una letra minúscula', 'error')
                return render_template('register.html')
            
            if not re.search(r'[0-9]', password):
                flash('La contraseña debe contener al menos un número', 'error')
                return render_template('register.html')
            
            if not re.search(r'[!@#$%^&*(),.?":{}|<>_\-]', password):
                flash('La contraseña debe contener al menos un carácter especial (!@#$%^&*...)', 'error')
                return render_template('register.html')
            
            check_query = "SELECT id FROM usuarios WHERE documento = %s"
            existing_user = db_manager.obtener_uno(check_query, (documento,))
            if existing_user:
                flash('El documento ya está registrado', 'error')
                return render_template('register.html')
            
            check_email = "SELECT id FROM usuarios WHERE email = %s"
            existing_email = db_manager.obtener_uno(check_email, (email,))
            if existing_email:
                flash('El email ya está registrado', 'error')
                return render_template('register.html')
            
            import bcrypt
            password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            
            insert_query = """
                INSERT INTO usuarios (documento, nombres, apellidos, email, telefono, password_hash, estado)
                VALUES (%s, %s, %s, %s, %s, %s, 'inactivo')
            """
            db_manager.ejecutar_comando(insert_query, (documento, nombres, apellidos, email, telefono, password_hash))
            
            try:
                log_query = """
                    INSERT INTO logs_sistema (modulo, nivel_log, mensaje, ip_address)
                    VALUES ('auth', 'INFO', %s, %s)
                """
                mensaje = f'Nuevo registro de usuario: {documento} - {nombres} {apellidos}'
                db_manager.ejecutar_comando(log_query, (mensaje, request.remote_addr))
            except:
                pass
            
            flash('Registro exitoso. Su cuenta será activada por un administrador en breve. Recibirá una notificación cuando pueda acceder.', 'success')
            return redirect(url_for('login'))
            
        except Exception as e:
            print(f"Error en registro: {e}")
            flash('Error al registrar usuario. Intente nuevamente.', 'error')
            return render_template('register.html')
    
    return render_template('register.html')

@app.route('/dashboard')
def dashboard():
    """Dashboard principal"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    stats = obtener_estadisticas_sistema()
    return render_template('dashboard.html', user=get_user_data(), stats=stats)

@app.route('/equipos')
def equipos():
    """Página de gestión de equipos"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    equipos_list = []
    categorias_list = []
    laboratorios_list = []
    
    try:
        # Obtener categorías
        cat_query = "SELECT id, nombre, codigo FROM categorias_equipos ORDER BY nombre"
        categorias_list = db_manager.ejecutar_query(cat_query) or []
        
        # Obtener laboratorios
        lab_query = "SELECT id, nombre, codigo_lab FROM laboratorios ORDER BY nombre"
        laboratorios_list = db_manager.ejecutar_query(lab_query) or []
        
        # Obtener equipos con toda la información
        query = """
            SELECT e.id, e.codigo_interno, e.codigo_qr, e.nombre, e.marca, e.modelo,
                   e.numero_serie, e.descripcion, e.especificaciones_tecnicas,
                   e.valor_adquisicion, e.proveedor, e.garantia_meses, e.vida_util_anos,
                   e.imagen_url, e.imagen_hash,
                   e.estado, e.estado_fisico, e.ubicacion_especifica, e.observaciones,
                   e.id_categoria, e.id_laboratorio,
                   DATE_FORMAT(e.fecha_adquisicion, '%d/%m/%Y') as fecha_adquisicion,
                   DATE_FORMAT(e.fecha_registro, '%d/%m/%Y') as fecha_registro,
                   l.nombre as laboratorio_nombre,
                   l.codigo_lab as laboratorio_codigo,
                   c.nombre as categoria_nombre,
                   c.codigo as categoria_codigo,
                   (SELECT DATE_FORMAT(MAX(fecha_inicio), '%d/%m/%Y') 
                    FROM historial_mantenimiento WHERE id_equipo = e.id AND estado = 'completado') as ultimo_mantenimiento,
                   (SELECT DATE_FORMAT(MIN(proxima_fecha_mantenimiento), '%d/%m/%Y') 
                    FROM historial_mantenimiento WHERE id_equipo = e.id AND proxima_fecha_mantenimiento >= CURDATE() AND estado = 'completado') as proximo_mantenimiento
            FROM equipos e
            LEFT JOIN laboratorios l ON e.id_laboratorio = l.id
            LEFT JOIN categorias_equipos c ON e.id_categoria = c.id
            ORDER BY e.nombre
        """
        resultados = db_manager.ejecutar_query(query) or []
        
        for item in resultados:
            equipos_list.append({
                'id': item.get('id'),
                'codigo_interno': item.get('codigo_interno'),
                'codigo_qr': item.get('codigo_qr') or '',
                'nombre': item.get('nombre'),
                'marca': item.get('marca') or '',
                'modelo': item.get('modelo') or '',
                'numero_serie': item.get('numero_serie') or '',
                'descripcion': item.get('descripcion') or '',
                'especificaciones': item.get('especificaciones_tecnicas') or '',
                'valor_adquisicion': float(item.get('valor_adquisicion') or 0),
                'proveedor': item.get('proveedor') or '',
                'garantia_meses': item.get('garantia_meses') or 12,
                'vida_util_anos': item.get('vida_util_anos') or 5,
                'imagen_url': item.get('imagen_url') or '',
                'imagen_hash': item.get('imagen_hash') or '',
                'estado': item.get('estado') or 'disponible',
                'estado_fisico': item.get('estado_fisico') or 'bueno',
                'ubicacion': item.get('ubicacion_especifica') or '',
                'observaciones': item.get('observaciones') or '',
                'fecha_adquisicion': item.get('fecha_adquisicion') or '',
                'fecha_registro': item.get('fecha_registro') or '',
                'id_laboratorio': item.get('id_laboratorio'),
                'laboratorio_nombre': item.get('laboratorio_nombre') or 'Sin asignar',
                'laboratorio_codigo': item.get('laboratorio_codigo') or '',
                'id_categoria': item.get('id_categoria'),
                'categoria': item.get('categoria_nombre') or 'General',
                'categoria_codigo': item.get('categoria_codigo') or '',
                'ultimo_mantenimiento': item.get('ultimo_mantenimiento') or 'N/A',
                'proximo_mantenimiento': item.get('proximo_mantenimiento') or 'N/A'
            })
    except Exception as e:
        print(f"Error obteniendo equipos: {e}")
        import traceback
        traceback.print_exc()
    
    return render_template('equipos.html', 
                           user=get_user_data(), 
                           equipos=equipos_list,
                           categorias=categorias_list,
                           laboratorios=laboratorios_list)

@app.route('/reconocimiento')
def reconocimiento_equipos_view():
    """Página de reconocimiento de equipos con IA"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    return render_template('reconocimiento_equipos.html', user=get_user_data())

@app.route('/laboratorios')
@app.route('/laboratorio')
@app.route('/laboratorio/<int:lab_id>')
def laboratorios(lab_id=None):
    """Página de gestión de laboratorios e inventarios"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    laboratorios_list = []
    instructores_list = []
    try:
        query = """
            SELECT l.id, l.codigo_lab as codigo, l.nombre, l.tipo, l.ubicacion,
                   l.capacidad_personas as capacidad_estudiantes, l.area_m2, l.estado,
                   l.responsable_id,
                   CONCAT(u.nombres, ' ', u.apellidos) as responsable,
                   (SELECT COUNT(*) FROM equipos WHERE id_laboratorio = l.id) as total_equipos,
                   0 as total_items,
                   0 as items_criticos
            FROM laboratorios l
            LEFT JOIN usuarios u ON l.responsable_id = u.id
            ORDER BY l.nombre
        """
        laboratorios_list = db_manager.ejecutar_query(query) or []
        
        # Obtener usuarios para el select de responsable
        query_instructores = """
            SELECT u.id, CONCAT(u.nombres, ' ', u.apellidos) as nombre, r.nombre_rol as especialidad
            FROM usuarios u
            LEFT JOIN roles r ON u.id_rol = r.id
            WHERE u.estado = 'activo'
            ORDER BY u.nombres
        """
        instructores_list = db_manager.ejecutar_query(query_instructores) or []
    except Exception as e:
        print(f"Error obteniendo laboratorios: {e}")
    
    return render_template('laboratorios.html', user=get_user_data(), laboratorios=laboratorios_list, instructores=instructores_list, lab_id=lab_id, puede_editar=puede_editar('laboratorios'))

@app.route('/inventario')
def inventario():
    """Página de gestión de inventario - Buscador Global"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Obtener laboratorios para el filtro
    laboratorios_list = []
    try:
        query_labs = "SELECT id, nombre FROM laboratorios ORDER BY nombre"
        laboratorios_list = db_manager.ejecutar_query(query_labs)
    except Exception as e:
        print(f"Error obteniendo laboratorios: {e}")
    
    # Obtener equipos
    equipos_list = []
    try:
        query_equipos = """
            SELECT e.id, e.nombre, e.estado, e.ubicacion_especifica as ubicacion,
                   c.nombre as tipo, l.id as laboratorio_id, l.nombre as laboratorio_nombre,
                   DATE_FORMAT(e.fecha_actualizacion, '%d/%m/%Y') as calibracion
            FROM equipos e
            LEFT JOIN categorias_equipos c ON e.id_categoria = c.id
            LEFT JOIN laboratorios l ON e.id_laboratorio = l.id
            ORDER BY e.nombre
        """
        equipos_list = db_manager.ejecutar_query(query_equipos)
    except Exception as e:
        print(f"Error obteniendo equipos: {e}")
    
    # Obtener datos de inventario
    inventario_list = []
    try:
        query = """
            SELECT i.id, i.codigo, i.nombre, i.descripcion as categoria, 
                   i.cantidad_actual, i.cantidad_minima, i.unidad_medida as unidad, 
                   i.proveedor, i.ubicacion, i.id_laboratorio as laboratorio_id,
                   l.nombre as laboratorio_nombre,
                   DATE_FORMAT(i.fecha_vencimiento, '%d/%m/%Y') as vencimiento
            FROM inventario i
            LEFT JOIN laboratorios l ON i.id_laboratorio = l.id
            ORDER BY i.nombre
        """
        resultados = db_manager.ejecutar_query(query)
        
        for item in resultados:
            cantidad_actual = item.get('cantidad_actual') or 0
            cantidad_minima = item.get('cantidad_minima') or 0
            
            # Calcular nivel de stock
            if cantidad_actual <= cantidad_minima:
                nivel_stock = 'critico'
            elif cantidad_actual <= cantidad_minima * 1.5:
                nivel_stock = 'bajo'
            else:
                nivel_stock = 'normal'
            
            inventario_list.append({
                'id': item.get('id'),
                'nombre': item.get('nombre'),
                'categoria': item.get('categoria') or 'General',
                'cantidad_actual': cantidad_actual,
                'cantidad_minima': cantidad_minima,
                'unidad': item.get('unidad'),
                'proveedor': item.get('proveedor'),
                'ubicacion': item.get('ubicacion'),
                'laboratorio_id': item.get('laboratorio_id'),
                'laboratorio_nombre': item.get('laboratorio_nombre') or 'Sin asignar',
                'vencimiento': item.get('vencimiento'),
                'nivel_stock': nivel_stock
            })
    except Exception as e:
        print(f"Error obteniendo inventario: {e}")
    
    return render_template('inventario.html', 
                           user=get_user_data(), 
                           inventario=inventario_list,
                           equipos=equipos_list,
                           laboratorios=laboratorios_list)

@app.route('/reservas')
def reservas():
    """Página de gestión de reservas"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Obtener préstamos de la base de datos (query adaptada a schema.sql)
    try:
        query = """
            SELECT 
                p.id,
                p.codigo,
                p.fecha as fecha_inicio,
                p.fecha_devolucion_programada as fecha_fin,
                p.estado,
                p.observaciones,
                p.proposito,
                e.nombre as equipo_nombre,
                e.codigo_interno as equipo_codigo,
                CONCAT(u.nombres, ' ', u.apellidos) as usuario_nombre,
                u.id as usuario_id
            FROM prestamos p
            LEFT JOIN equipos e ON p.id_equipo = e.id
            LEFT JOIN usuarios u ON p.id_usuario_solicitante = u.id
            ORDER BY p.fecha_solicitud DESC
            LIMIT 100
        """
        reservas_list = db_manager.ejecutar_query(query) or []
        print(f"✅ Se obtuvieron {len(reservas_list)} préstamos de la base de datos")
    except Exception as e:
        print(f"❌ Error obteniendo préstamos: {e}")
        import traceback
        traceback.print_exc()
        reservas_list = []
    
    return render_template('reservas.html', user=get_user_data(), reservas=reservas_list)

@app.route('/prestamos')
def prestamos():
    """Página de gestión de préstamos de equipos"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar permisos: prestamos (gestión completa) o prestamos_propios (solo propios)
    permisos = session.get('user_permisos', {})
    tiene_acceso = permisos.get('all') or permisos.get('prestamos') or permisos.get('prestamos_propios')
    
    if not tiene_acceso:
        flash('No tiene permisos para acceder a préstamos.', 'error')
        return redirect(url_for('dashboard'))
    
    # puede_gestionar: aprobar, rechazar, activar préstamos de otros
    puede_gestionar = permisos.get('all') or permisos.get('prestamos')
    # solo_propios: solo ve y solicita sus propios préstamos
    solo_propios = not puede_gestionar and permisos.get('prestamos_propios')
    
    return render_template('prestamos.html', 
                           user=get_user_data(), 
                           puede_editar=puede_gestionar,
                           solo_propios=solo_propios)

@app.route('/registro-facial')
def registro_facial():
    """Página de registro de reconocimiento facial"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_data = {
        'user_id': session.get('user_id'),
        'user_name': session.get('user_name', 'Usuario'),
        'user_level': session.get('user_level', 1)
    }
    
    return render_template('registro_facial.html', user=user_data)

@app.route('/registro-equipos-ia')
def registro_equipos_ia():
    """Página de registro de equipos con IA - Objetivo MGA"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('ia_visual'):
        flash('No tiene permisos para registrar equipos con IA', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('registro_equipos_ia.html', user=get_user_data())

@app.route('/registros-gestion')
def registros_gestion():
    """Página de gestión de registros de equipos"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('ia_visual'):
        flash('No tiene permisos para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('registros_gestion.html', user=get_user_data())

@app.route('/entrenamiento-ia')
def entrenamiento_ia():
    """Página de entrenamiento del modelo MobileNet"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('ia_visual'):
        flash('No tiene permisos para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('entrenamiento_visual.html', user=get_user_data())

# ========================================
# RUTAS DE GESTIÓN ADICIONALES
# ========================================

@app.route('/roles')
def roles():
    """Página de gestión de roles"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('roles'):
        flash('No tiene permisos para gestionar roles.', 'error')
        return redirect(url_for('dashboard'))
    
    roles_lista = []
    try:
        query = "SELECT id, nombre_rol, descripcion, permisos, estado FROM roles ORDER BY id"
        roles_lista = db_manager.ejecutar_query(query) or []
    except Exception as e:
        flash(f'Error cargando roles: {str(e)}', 'error')
    
    return render_template('roles.html', user=get_user_data(), roles=roles_lista, puede_editar=puede_editar('roles'))

@app.route('/programas')
def programas():
    """Página de gestión de programas de formación"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('programas'):
        flash('No tiene permisos para gestionar programas.', 'error')
        return redirect(url_for('dashboard'))
    
    programas_lista = []
    try:
        query = """
            SELECT id, codigo_programa, nombre_programa, tipo_programa, 
                   descripcion, duracion_meses, estado
            FROM programas_formacion 
            ORDER BY nombre_programa
        """
        programas_lista = db_manager.ejecutar_query(query) or []
    except Exception as e:
        flash(f'Error cargando programas: {str(e)}', 'error')
    
    return render_template('programas.html', user=get_user_data(), programas=programas_lista, puede_editar=puede_editar('programas'))

@app.route('/practicas')
def practicas():
    """Página de gestión de prácticas de laboratorio"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('reservas'):
        flash('No tiene permisos para gestionar reservas.', 'error')
        return redirect(url_for('dashboard'))
    
    practicas_lista = []
    laboratorios_lista = []
    programas_lista = []
    instructores_lista = []
    
    try:
        db_manager.ejecutar_comando("""
            UPDATE practicas_laboratorio
            SET estado = 'completada'
            WHERE estado IN ('programada', 'en_curso')
            AND DATE_ADD(fecha, INTERVAL ROUND(COALESCE(duracion_horas, 1) * 60) MINUTE) <= NOW()
        """)

        db_manager.ejecutar_comando("""
            UPDATE practicas_laboratorio
            SET estado = 'en_curso'
            WHERE estado = 'programada'
            AND fecha <= NOW()
            AND DATE_ADD(fecha, INTERVAL ROUND(COALESCE(duracion_horas, 1) * 60) MINUTE) > NOW()
        """)

        # Obtener ID del instructor si el usuario es instructor
        user_rol = session.get('user_rol')
        user_id = session.get('user_id')
        instructor_id = None
        es_instructor = False
        
        if user_rol == 2:  # Rol de Instructor
            es_instructor = True
            query_instructor = "SELECT id FROM instructores WHERE id_usuario = %s"
            resultado = db_manager.ejecutar_query(query_instructor, (user_id,))
            if resultado and len(resultado) > 0:
                instructor_id = resultado[0]['id']
        
        # Obtener prácticas con joins (filtrar por instructor si aplica)
        query = """
            SELECT p.id, p.codigo, p.nombre, p.id_programa, p.id_laboratorio, 
                   p.id_instructor, p.fecha, p.duracion_horas, p.numero_estudiantes,
                   p.equipos_requeridos, p.materiales_requeridos, p.objetivos,
                   p.descripcion_actividades, p.observaciones, p.estado,
                   DATE_FORMAT(p.fecha, '%d/%m/%Y %H:%i') as fecha_formato,
                   pf.nombre_programa,
                   l.nombre as laboratorio_nombre,
                   CONCAT(u.nombres, ' ', u.apellidos) as instructor_nombre,
                   (SELECT COUNT(*)
                    FROM prestamos pr
                    WHERE pr.codigo LIKE CONCAT('PRAC-', p.codigo, '-%')) as prestamos_total,
                   (SELECT COUNT(*)
                    FROM prestamos pr
                    WHERE pr.codigo LIKE CONCAT('PRAC-', p.codigo, '-%')
                    AND pr.estado = 'activo') as prestamos_activos
            FROM practicas_laboratorio p
            LEFT JOIN programas_formacion pf ON p.id_programa = pf.id
            LEFT JOIN laboratorios l ON p.id_laboratorio = l.id
            LEFT JOIN instructores i ON p.id_instructor = i.id
            LEFT JOIN usuarios u ON i.id_usuario = u.id
        """
        
        # Si es instructor, filtrar solo sus prácticas
        if es_instructor and instructor_id:
            query += " WHERE p.id_instructor = %s"
            practicas_lista = db_manager.ejecutar_query(query + " ORDER BY p.fecha DESC", (instructor_id,)) or []
        else:
            practicas_lista = db_manager.ejecutar_query(query + " ORDER BY p.fecha DESC") or []
        
        # Obtener laboratorios para filtros y formulario
        query_labs = "SELECT id, nombre FROM laboratorios WHERE estado = 'disponible' ORDER BY nombre"
        laboratorios_lista = db_manager.ejecutar_query(query_labs) or []
        
        # Obtener programas para filtros y formulario
        query_progs = "SELECT id, nombre_programa FROM programas_formacion WHERE estado = 'activo' ORDER BY nombre_programa"
        programas_lista = db_manager.ejecutar_query(query_progs) or []
        
        # Obtener instructores para formulario
        query_inst = """
            SELECT i.id, CONCAT(u.nombres, ' ', u.apellidos) as nombre_completo
            FROM instructores i
            JOIN usuarios u ON i.id_usuario = u.id
            WHERE u.estado = 'activo'
            ORDER BY u.nombres
        """
        instructores_lista = db_manager.ejecutar_query(query_inst) or []
        
    except Exception as e:
        print(f"Error obteniendo datos de prácticas: {str(e)}")
    
    return render_template('practicas.html', 
                           user=get_user_data(), 
                           practicas=practicas_lista,
                           laboratorios=laboratorios_lista,
                           programas=programas_lista,
                           instructores=instructores_lista,
                           puede_editar=puede_editar('reservas'),
                           es_instructor=es_instructor,
                           instructor_id=instructor_id)

@app.route('/estadisticas-practicas')
def estadisticas_practicas():
    """Página de estadísticas y monitoreo de prácticas"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('reservas'):
        flash('No tiene permisos para ver estadísticas de reservas.', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('estadisticas_practicas.html', user=get_user_data())

@app.route('/mantenimiento')
def mantenimiento():
    """Página de gestión de mantenimiento"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('mantenimiento'):
        flash('No tiene permisos para acceder a mantenimiento.', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('mantenimiento.html', user=get_user_data(), puede_editar=puede_editar('mantenimiento'))

@app.route('/capacitaciones', methods=['GET', 'POST'])
def capacitaciones():
    """Página de gestión de capacitaciones - Programa Formativo en IA"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('capacitaciones'):
        flash('No tiene permisos para acceder a capacitaciones.', 'error')
        return redirect(url_for('dashboard'))
    
    # Manejar POST para crear/editar capacitación
    if request.method == 'POST':
        print(f"POST recibido. puede_editar: {puede_editar('capacitaciones')}")

        action = request.form.get('action', 'crear')
        
        if not puede_editar('capacitaciones'):
            flash('No tiene permisos para gestionar capacitaciones', 'error')
            return redirect(url_for('capacitaciones'))
        
        # Editar capacitación existente
        if action == 'editar':
            try:
                cap_id = request.form.get('id')
                titulo = request.form.get('titulo', '').strip()
                descripcion = request.form.get('descripcion', '').strip()
                tipo_capacitacion = request.form.get('tipo_capacitacion')
                estado = request.form.get('estado')
                producto = request.form.get('producto', '').strip()
                medicion = request.form.get('medicion', '').strip()
                
                # Convertir campos numéricos vacíos a 0
                cantidad_meta = request.form.get('cantidad_meta', '0').strip()
                cantidad_meta = int(cantidad_meta) if cantidad_meta else 0
                
                cantidad_actual = request.form.get('cantidad_actual', '0').strip()
                cantidad_actual = int(cantidad_actual) if cantidad_actual else 0
                
                actividad = request.form.get('actividad', '').strip()
                duracion_horas = request.form.get('duracion_horas')
                fecha_inicio = request.form.get('fecha_inicio')
                fecha_fin = request.form.get('fecha_fin')
                
                # Calcular porcentaje de avance
                porcentaje_avance = 0
                if cantidad_meta and int(cantidad_meta) > 0:
                    porcentaje_avance = (int(cantidad_actual) / int(cantidad_meta)) * 100
                
                query_update = """
                    UPDATE capacitaciones SET
                        titulo = %s, descripcion = %s, tipo_capacitacion = %s,
                        producto = %s, medicion = %s, cantidad_meta = %s,
                        cantidad_actual = %s, actividad = %s, porcentaje_avance = %s,
                        duracion_horas = %s, fecha_inicio = %s, fecha_fin = %s, estado = %s
                    WHERE id = %s
                """
                
                db_manager.ejecutar_comando(query_update, (
                    titulo, descripcion, tipo_capacitacion, producto, medicion,
                    cantidad_meta, cantidad_actual, actividad, porcentaje_avance,
                    duracion_horas, fecha_inicio, fecha_fin, estado, cap_id
                ))
                
                flash('Capacitación actualizada exitosamente', 'success')
                return redirect(url_for('capacitaciones'))
                
            except Exception as e:
                flash(f'Error al actualizar capacitación: {str(e)}', 'error')
                print(f"Error actualizando capacitación: {str(e)}")
                import traceback
                traceback.print_exc()
                return redirect(url_for('capacitaciones'))
        
        # Crear nueva capacitación
        else:
            try:
                print("=== INICIANDO CREACIÓN DE CAPACITACIÓN ===")
                titulo = request.form.get('titulo', '').strip()
                descripcion = request.form.get('descripcion', '').strip()
                tipo_capacitacion = request.form.get('tipo_capacitacion')
                estado = request.form.get('estado', 'programada')
                producto = request.form.get('producto', '').strip()
                medicion = request.form.get('medicion', '').strip()
                
                # Convertir campos numéricos vacíos a 0
                cantidad_meta = request.form.get('cantidad_meta', '0').strip()
                cantidad_meta = int(cantidad_meta) if cantidad_meta else 0
                
                cantidad_actual = request.form.get('cantidad_actual', '0').strip()
                cantidad_actual = int(cantidad_actual) if cantidad_actual else 0
                
                actividad = request.form.get('actividad', '').strip()
                duracion_horas = request.form.get('duracion_horas')
                fecha_inicio = request.form.get('fecha_inicio')
                fecha_fin = request.form.get('fecha_fin')
                
                print(f"Título: {titulo}")
                print(f"Tipo: {tipo_capacitacion}")
                print(f"Estado: {estado}")
                print(f"Duración: {duracion_horas}")
                print(f"Fechas: {fecha_inicio} - {fecha_fin}")
                
                # Validaciones backend
                if not titulo or len(titulo) < 5 or len(titulo) > 300:
                    print("❌ Error: Título inválido")
                    print(f"Error: Título inválido: {titulo}")
                    flash('El título debe tener entre 5 y 300 caracteres', 'error')
                    return redirect(url_for('capacitaciones'))
                
                if not tipo_capacitacion or tipo_capacitacion not in ['modulo_formativo', 'taller', 'material_didactico', 'gestion_cambio']:
                    print(f"❌ Error: Tipo inválido: {tipo_capacitacion}")
                    flash('Tipo de capacitación inválido', 'error')
                    return redirect(url_for('capacitaciones'))
                
                if not duracion_horas or int(duracion_horas) < 1 or int(duracion_horas) > 500:
                    print(f"❌ Error: Duración inválida: {duracion_horas}")
                    flash('La duración debe estar entre 1 y 500 horas', 'error')
                    return redirect(url_for('capacitaciones'))
                
                if not fecha_inicio or not fecha_fin:
                    print(f"❌ Error: Fechas faltantes: {fecha_inicio} - {fecha_fin}")
                    flash('Las fechas de inicio y fin son requeridas', 'error')
                    return redirect(url_for('capacitaciones'))
                
                # Validar que fecha_fin sea posterior a fecha_inicio
                from datetime import datetime
                fecha_i = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                fecha_f = datetime.strptime(fecha_fin, '%Y-%m-%d')
                if fecha_f < fecha_i:
                    flash('La fecha de fin debe ser posterior a la fecha de inicio', 'error')
                    return redirect(url_for('capacitaciones'))
                
                # Validar cantidad actual vs meta
                if cantidad_meta and cantidad_actual:
                    if int(cantidad_actual) > int(cantidad_meta):
                        flash('La cantidad actual no puede ser mayor que la meta', 'error')
                        return redirect(url_for('capacitaciones'))
                
                # Calcular porcentaje de avance
                porcentaje_avance = 0
                if cantidad_meta and int(cantidad_meta) > 0:
                    porcentaje_avance = (int(cantidad_actual) / int(cantidad_meta)) * 100
                
                print("✅ Todas las validaciones pasaron")
                print(f"Insertando capacitación: {titulo}")
                
                query_insert = """
                    INSERT INTO capacitaciones (
                        titulo, descripcion, tipo_capacitacion, producto, medicion,
                        cantidad_meta, cantidad_actual, actividad, porcentaje_avance,
                        duracion_horas, fecha_inicio, fecha_fin, estado
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                db_manager.ejecutar_comando(query_insert, (
                    titulo, descripcion, tipo_capacitacion, producto, medicion,
                    cantidad_meta, cantidad_actual, actividad, porcentaje_avance,
                    duracion_horas, fecha_inicio, fecha_fin, estado
                ))
                
                print("✅ Capacitación insertada en BD")
                flash('Capacitación creada exitosamente', 'success')
                print("✅ Flash message agregado")
                return redirect(url_for('capacitaciones'))
                
            except Exception as e:
                flash(f'Error al crear capacitación: {str(e)}', 'error')
                print(f"Error creando capacitación: {str(e)}")
                import traceback
                traceback.print_exc()
    
    # Cargar lista de capacitaciones
    capacitaciones_lista = []
    try:
        query = """
            SELECT c.id, c.titulo, c.descripcion, c.tipo_capacitacion, 
                   c.producto, c.medicion, c.cantidad_meta, c.cantidad_actual,
                   c.actividad, c.porcentaje_avance, c.duracion_horas, c.estado,
                   DATE_FORMAT(c.fecha_inicio, '%d/%m/%Y') as fecha_inicio,
                   DATE_FORMAT(c.fecha_fin, '%d/%m/%Y') as fecha_fin,
                   CONCAT(u.nombres, ' ', u.apellidos) as instructor
            FROM capacitaciones c
            LEFT JOIN usuarios u ON c.id_instructor = u.id
            ORDER BY 
                FIELD(c.tipo_capacitacion, 'modulo_formativo', 'taller', 'material_didactico', 'gestion_cambio'),
                c.fecha_inicio ASC
        """
        capacitaciones_lista = db_manager.ejecutar_query(query) or []
    except Exception as e:
        print(f"Error cargando capacitaciones: {str(e)}")
        import traceback
        traceback.print_exc()
    
    return render_template('capacitaciones.html', user=get_user_data(), capacitaciones=capacitaciones_lista, puede_editar=puede_editar('capacitaciones'))

@app.route('/asistente')
def asistente_lucia():
    """Página del Asistente de Voz LUCIA"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    return render_template('asistente_lucia.html', user=get_user_data())

@app.route('/usuarios')
def usuarios():
    """Página de gestión de usuarios"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('usuarios', incluir_ver=False):
        flash('No tiene permisos para acceder a esta página.', 'error')
        return redirect(url_for('dashboard'))
    
    user_data = get_user_data()
    es_admin = puede_editar('usuarios')
    
    NIVELES_ROL = {1: 5, 2: 4, 3: 3, 4: 2, 5: 2}
    
    usuarios_lista = []
    roles_lista = []
    try:
        # Cargar roles para los selects
        query_roles = "SELECT id, nombre_rol FROM roles WHERE estado = 'activo' ORDER BY id"
        roles_lista = db_manager.ejecutar_query(query_roles) or []
        
        query = """
            SELECT u.id, u.documento, u.nombres, u.apellidos, u.email, u.telefono,
                   u.id_rol, u.estado,
                   DATE_FORMAT(u.fecha_registro, '%d/%m/%Y') as fecha_registro,
                   r.nombre_rol
            FROM usuarios u
            LEFT JOIN roles r ON u.id_rol = r.id
            ORDER BY u.nombres, u.apellidos
        """
        usuarios_db = db_manager.ejecutar_query(query)
        
        for u in usuarios_db:
            usuarios_lista.append({
                'id': u['id'],
                'documento': u['documento'],
                'nombre': f"{u['nombres']} {u['apellidos']}",
                'rol': u['nombre_rol'] or 'Usuario',
                'id_rol': u['id_rol'],
                'tipo': u['nombre_rol'] or 'usuario',
                'nivel_acceso': NIVELES_ROL.get(u['id_rol'], 1),
                'activo': u['estado'] == 'activo',
                'email': u['email'],
                'telefono': u['telefono'],
                'registro': u['fecha_registro'],
                'tiene_rostro': 'No'
            })
    except Exception as e:
        flash(f'Error cargando usuarios: {str(e)}', 'error')
    
    return render_template('usuarios.html', user=user_data, usuarios=usuarios_lista, roles=roles_lista, es_admin=es_admin)

@app.route('/reportes')
def reportes():
    """Página de reportes y estadísticas"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('reportes'):
        flash('No tiene permisos para acceder a reportes.', 'error')
        return redirect(url_for('dashboard'))
    
    from datetime import datetime, timedelta
    
    # Obtener filtros de fecha y tipo
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    tipo_reporte = request.args.get('tipo_reporte', 'general')
    
    # Si no hay fechas, usar último mes
    if not fecha_inicio or not fecha_fin:
        fecha_fin = datetime.now().strftime('%Y-%m-%d')
        fecha_inicio = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    
    try:
        # ===== ESTADÍSTICAS GENERALES =====
        stats = {}
        
        # Total de equipos
        query_equipos = "SELECT COUNT(*) as total FROM equipos WHERE estado != 'dado_baja'"
        result = db_manager.ejecutar_query(query_equipos)
        stats['total_equipos'] = result[0]['total'] if result else 0
        
        # Equipos disponibles
        query_disponibles = "SELECT COUNT(*) as total FROM equipos WHERE estado = 'disponible'"
        result = db_manager.ejecutar_query(query_disponibles)
        stats['equipos_disponibles'] = result[0]['total'] if result else 0
        
        # Préstamos activos
        query_prestamos = "SELECT COUNT(*) as total FROM prestamos WHERE estado = 'activo'"
        result = db_manager.ejecutar_query(query_prestamos)
        stats['prestamos_activos'] = result[0]['total'] if result else 0
        
        # Préstamos vencidos
        query_vencidos = """
            SELECT COUNT(*) as total FROM prestamos 
            WHERE estado = 'activo' AND fecha_devolucion_programada < NOW()
        """
        result = db_manager.ejecutar_query(query_vencidos)
        stats['prestamos_vencidos'] = result[0]['total'] if result else 0
        
        # Mantenimientos del mes
        query_mant = """
            SELECT COUNT(*) as total FROM historial_mantenimiento 
            WHERE MONTH(fecha_inicio) = MONTH(NOW()) AND YEAR(fecha_inicio) = YEAR(NOW())
        """
        result = db_manager.ejecutar_query(query_mant)
        stats['mantenimientos_mes'] = result[0]['total'] if result else 0
        
        # Prácticas del mes
        query_prac = """
            SELECT COUNT(*) as total FROM practicas_laboratorio 
            WHERE MONTH(fecha) = MONTH(NOW()) AND YEAR(fecha) = YEAR(NOW())
        """
        result = db_manager.ejecutar_query(query_prac)
        stats['practicas_mes'] = result[0]['total'] if result else 0
        
        # ===== DATOS PARA TABLAS (según tipo_reporte) =====
        equipos = []
        prestamos = []
        mantenimientos = []
        practicas = []
        
        # Consultar solo los datos relevantes según el tipo de reporte
        if tipo_reporte == 'general' or tipo_reporte == 'equipos':
            # Equipos
            query_equipos_list = """
                SELECT e.codigo_interno, e.nombre, c.nombre as categoria, 
                       l.nombre as laboratorio, e.estado, e.estado_fisico, e.valor_adquisicion
                FROM equipos e
                LEFT JOIN categorias_equipos c ON e.id_categoria = c.id
                LEFT JOIN laboratorios l ON e.id_laboratorio = l.id
                WHERE e.estado != 'dado_baja'
                ORDER BY e.fecha_registro DESC
                LIMIT 100
            """
            equipos = db_manager.ejecutar_query(query_equipos_list) or []
        
        if tipo_reporte == 'general' or tipo_reporte == 'prestamos':
            # Préstamos
            query_prestamos_list = """
                SELECT p.codigo, e.nombre as equipo_nombre, 
                       CONCAT(u.nombres, ' ', u.apellidos) as solicitante,
                       DATE_FORMAT(p.fecha, '%d/%m/%Y') as fecha_prestamo,
                       DATE_FORMAT(p.fecha_devolucion_programada, '%d/%m/%Y') as fecha_devolucion_programada,
                       p.estado,
                       DATEDIFF(COALESCE(p.fecha_devolucion_real, NOW()), p.fecha) as dias_prestamo
                FROM prestamos p
                JOIN equipos e ON p.id_equipo = e.id
                JOIN usuarios u ON p.id_usuario_solicitante = u.id
                WHERE p.fecha BETWEEN %s AND %s
                ORDER BY p.fecha DESC
                LIMIT 100
            """
            prestamos = db_manager.ejecutar_query(query_prestamos_list, (fecha_inicio, fecha_fin)) or []
        
        if tipo_reporte == 'general' or tipo_reporte == 'mantenimiento':
            # Mantenimientos
            query_mant_list = """
                SELECT e.nombre as equipo_nombre, tm.nombre as tipo_mantenimiento,
                       CONCAT(u.nombres, ' ', u.apellidos) as tecnico,
                       DATE_FORMAT(hm.fecha_inicio, '%d/%m/%Y') as fecha_inicio,
                       DATE_FORMAT(hm.fecha_fin, '%d/%m/%Y') as fecha_fin,
                       hm.estado, hm.costo_mantenimiento
                FROM historial_mantenimiento hm
                JOIN equipos e ON hm.id_equipo = e.id
                JOIN tipos_mantenimiento tm ON hm.id_tipo_mantenimiento = tm.id
                LEFT JOIN usuarios u ON hm.tecnico_responsable_id = u.id
                WHERE hm.fecha_inicio BETWEEN %s AND %s
                ORDER BY hm.fecha_inicio DESC
                LIMIT 100
            """
            mantenimientos = db_manager.ejecutar_query(query_mant_list, (fecha_inicio, fecha_fin)) or []
        
        if tipo_reporte == 'general' or tipo_reporte == 'practicas':
            # Prácticas
            query_prac_list = """
                SELECT pl.codigo, pl.nombre, pf.nombre_programa as programa,
                       CONCAT(u.nombres, ' ', u.apellidos) as instructor,
                       l.nombre as laboratorio,
                       DATE_FORMAT(pl.fecha, '%d/%m/%Y %H:%i') as fecha,
                       pl.numero_estudiantes, pl.estado
                FROM practicas_laboratorio pl
                LEFT JOIN programas_formacion pf ON pl.id_programa = pf.id
                LEFT JOIN instructores i ON pl.id_instructor = i.id
                LEFT JOIN usuarios u ON i.id_usuario = u.id
                LEFT JOIN laboratorios l ON pl.id_laboratorio = l.id
                WHERE pl.fecha BETWEEN %s AND %s
                ORDER BY pl.fecha DESC
                LIMIT 100
            """
            practicas = db_manager.ejecutar_query(query_prac_list, (fecha_inicio, fecha_fin)) or []
        
        # ===== DATOS PARA GRÁFICOS =====
        datos_graficos = {}
        
        # Gráfico: Equipos por Estado
        query_eq_estado = """
            SELECT estado, COUNT(*) as total 
            FROM equipos 
            WHERE estado != 'dado_baja'
            GROUP BY estado
        """
        result = db_manager.ejecutar_query(query_eq_estado) or []
        datos_graficos['equipos_estado'] = {
            'labels': [r['estado'].capitalize() for r in result],
            'data': [r['total'] for r in result]
        }
        
        # Gráfico: Préstamos por Mes (últimos 6 meses)
        query_prest_mes = """
            SELECT DATE_FORMAT(fecha, '%Y-%m') as mes, COUNT(*) as total
            FROM prestamos
            WHERE fecha >= DATE_SUB(NOW(), INTERVAL 6 MONTH)
            GROUP BY mes
            ORDER BY mes
        """
        result = db_manager.ejecutar_query(query_prest_mes) or []
        datos_graficos['prestamos_mes'] = {
            'labels': [r['mes'] for r in result],
            'data': [r['total'] for r in result]
        }
        
        # Gráfico: Mantenimientos por Tipo
        query_mant_tipo = """
            SELECT tm.nombre, COUNT(*) as total
            FROM historial_mantenimiento hm
            JOIN tipos_mantenimiento tm ON hm.id_tipo_mantenimiento = tm.id
            WHERE hm.fecha_inicio >= DATE_SUB(NOW(), INTERVAL 3 MONTH)
            GROUP BY tm.nombre
        """
        result = db_manager.ejecutar_query(query_mant_tipo) or []
        datos_graficos['mantenimientos'] = {
            'labels': [r['nombre'] for r in result],
            'data': [r['total'] for r in result]
        }
        
        # Gráfico: Uso de Laboratorios (horas de prácticas)
        query_lab_uso = """
            SELECT l.nombre, COALESCE(SUM(pl.duracion_horas), 0) as horas_uso
            FROM laboratorios l
            LEFT JOIN practicas_laboratorio pl ON l.id = pl.id_laboratorio
                AND pl.fecha >= DATE_SUB(NOW(), INTERVAL 1 MONTH)
            GROUP BY l.id, l.nombre
            ORDER BY horas_uso DESC
            LIMIT 5
        """
        result = db_manager.ejecutar_query(query_lab_uso) or []
        datos_graficos['laboratorios'] = {
            'labels': [r['nombre'] for r in result],
            'data': [float(r['horas_uso']) for r in result]
        }
        
    except Exception as e:
        print(f"Error cargando reportes: {str(e)}")
        import traceback
        traceback.print_exc()
        flash('Error cargando datos de reportes', 'error')
        stats = {
            'total_equipos': 0, 'equipos_disponibles': 0, 
            'prestamos_activos': 0, 'prestamos_vencidos': 0,
            'mantenimientos_mes': 0, 'practicas_mes': 0
        }
        equipos = []
        prestamos = []
        mantenimientos = []
        practicas = []
        datos_graficos = {
            'equipos_estado': {'labels': [], 'data': []},
            'prestamos_mes': {'labels': [], 'data': []},
            'mantenimientos': {'labels': [], 'data': []},
            'laboratorios': {'labels': [], 'data': []},
            'practicas': {'labels': [], 'data': []}
        }
    
    return render_template('reportes.html', 
                          user=get_user_data(),
                          stats=stats,
                          equipos=equipos,
                          prestamos=prestamos,
                          mantenimientos=mantenimientos,
                          practicas=practicas,
                          datos_graficos=datos_graficos,
                          fecha_inicio=fecha_inicio,
                          fecha_fin=fecha_fin,
                          tipo_reporte=tipo_reporte)

@app.route('/api/reportes/exportar/pdf')
def exportar_reporte_pdf():
    """Exportar reporte a PDF"""
    if 'user_id' not in session:
        return {'success': False, 'message': 'No autorizado'}, 401
    
    if not tiene_permiso('reportes'):
        return {'success': False, 'message': 'Sin permisos'}, 403
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from io import BytesIO
        from datetime import datetime
        
        # Obtener parámetros
        fecha_inicio = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        tipo_reporte = request.args.get('tipo_reporte', 'general')
        
        # Crear buffer para PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Contenedor de elementos
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        # Título
        elements.append(Paragraph("Reporte de Gestión de Laboratorios", title_style))
        elements.append(Paragraph(f"Centro Minero - SENA", styles['Normal']))
        elements.append(Paragraph(f"Período: {fecha_inicio} a {fecha_fin}", styles['Normal']))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Obtener datos (reutilizar lógica de reportes)
        from datetime import timedelta
        if not fecha_inicio or not fecha_fin:
            fecha_fin_dt = datetime.now().strftime('%Y-%m-%d')
            fecha_inicio_dt = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        else:
            fecha_inicio_dt = fecha_inicio
            fecha_fin_dt = fecha_fin
        
        # Estadísticas
        stats = {}
        query_equipos = "SELECT COUNT(*) as total FROM equipos WHERE estado != 'dado_baja'"
        result = db_manager.ejecutar_query(query_equipos)
        stats['total_equipos'] = result[0]['total'] if result else 0
        
        query_prestamos = "SELECT COUNT(*) as total FROM prestamos WHERE estado = 'activo'"
        result = db_manager.ejecutar_query(query_prestamos)
        stats['prestamos_activos'] = result[0]['total'] if result else 0
        
        # Tabla de estadísticas
        elements.append(Paragraph("Resumen Ejecutivo", styles['Heading2']))
        data_stats = [
            ['Métrica', 'Valor'],
            ['Total Equipos', str(stats['total_equipos'])],
            ['Préstamos Activos', str(stats['prestamos_activos'])]
        ]
        
        table_stats = Table(data_stats, colWidths=[3*inch, 2*inch])
        table_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table_stats)
        elements.append(Spacer(1, 20))
        
        # Agregar datos según el tipo de reporte
        if tipo_reporte == 'general' or tipo_reporte == 'equipos':
            # Equipos
            elements.append(Paragraph("Equipos", styles['Heading2']))
            query_equipos = """
                SELECT e.codigo_interno, e.nombre, c.nombre as categoria, 
                       l.nombre as laboratorio, e.estado
                FROM equipos e
                LEFT JOIN categorias_equipos c ON e.id_categoria = c.id
                LEFT JOIN laboratorios l ON e.id_laboratorio = l.id
                WHERE e.estado != 'dado_baja'
                ORDER BY e.fecha_registro DESC
                LIMIT 20
            """
            equipos = db_manager.ejecutar_query(query_equipos) or []
            
            if equipos:
                data_equipos = [['Código', 'Nombre', 'Categoría', 'Laboratorio', 'Estado']]
                for eq in equipos:
                    data_equipos.append([
                        eq['codigo_interno'] or 'N/A',
                        eq['nombre'][:30] if eq['nombre'] else 'N/A',
                        eq['categoria'][:20] if eq['categoria'] else 'N/A',
                        eq['laboratorio'][:20] if eq['laboratorio'] else 'N/A',
                        eq['estado']
                    ])
                
                table_equipos = Table(data_equipos, colWidths=[1*inch, 2*inch, 1.5*inch, 1.5*inch, 1*inch])
                table_equipos.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table_equipos)
                elements.append(Spacer(1, 20))
        
        if tipo_reporte == 'general' or tipo_reporte == 'prestamos':
            # Préstamos
            elements.append(Paragraph("Préstamos Recientes", styles['Heading2']))
            query_prestamos_list = """
                SELECT p.codigo, e.nombre as equipo, 
                       CONCAT(u.nombres, ' ', u.apellidos) as solicitante,
                       DATE_FORMAT(p.fecha, '%d/%m/%Y') as fecha,
                       p.estado
                FROM prestamos p
                JOIN equipos e ON p.id_equipo = e.id
                JOIN usuarios u ON p.id_usuario_solicitante = u.id
                WHERE p.fecha BETWEEN %s AND %s
                ORDER BY p.fecha DESC
                LIMIT 20
            """
            prestamos = db_manager.ejecutar_query(query_prestamos_list, (fecha_inicio_dt, fecha_fin_dt)) or []
            
            if prestamos:
                data_prestamos = [['Código', 'Equipo', 'Solicitante', 'Fecha', 'Estado']]
                for p in prestamos:
                    data_prestamos.append([
                        p['codigo'], 
                        p['equipo'][:30], 
                        p['solicitante'][:25],
                        p['fecha'],
                        p['estado']
                    ])
                
                table_prestamos = Table(data_prestamos, colWidths=[1*inch, 2*inch, 1.5*inch, 1*inch, 1*inch])
                table_prestamos.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table_prestamos)
                elements.append(Spacer(1, 20))
        
        if tipo_reporte == 'general' or tipo_reporte == 'mantenimiento':
            # Mantenimientos
            elements.append(Paragraph("Mantenimientos", styles['Heading2']))
            query_mant = """
                SELECT e.nombre as equipo, tm.nombre as tipo,
                       DATE_FORMAT(hm.fecha_inicio, '%d/%m/%Y') as fecha,
                       hm.estado
                FROM historial_mantenimiento hm
                JOIN equipos e ON hm.id_equipo = e.id
                JOIN tipos_mantenimiento tm ON hm.id_tipo_mantenimiento = tm.id
                WHERE hm.fecha_inicio BETWEEN %s AND %s
                ORDER BY hm.fecha_inicio DESC
                LIMIT 20
            """
            mantenimientos = db_manager.ejecutar_query(query_mant, (fecha_inicio_dt, fecha_fin_dt)) or []
            
            if mantenimientos:
                data_mant = [['Equipo', 'Tipo', 'Fecha', 'Estado']]
                for m in mantenimientos:
                    data_mant.append([
                        m['equipo'][:30],
                        m['tipo'][:20],
                        m['fecha'],
                        m['estado']
                    ])
                
                table_mant = Table(data_mant, colWidths=[2.5*inch, 2*inch, 1.5*inch, 1*inch])
                table_mant.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(table_mant)
        
        # Generar PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'reporte_laboratorios_{fecha_inicio}_{fecha_fin}.pdf'
        )
        
    except Exception as e:
        print(f"Error generando PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': f'Error: {str(e)}'}, 500

@app.route('/api/reportes/exportar/excel')
def exportar_reporte_excel():
    """Exportar reporte a Excel"""
    if 'user_id' not in session:
        return {'success': False, 'message': 'No autorizado'}, 401
    
    if not tiene_permiso('reportes'):
        return {'success': False, 'message': 'Sin permisos'}, 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from datetime import datetime, timedelta
        
        # Obtener parámetros
        fecha_inicio = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        
        if not fecha_inicio or not fecha_fin:
            fecha_fin = datetime.now().strftime('%Y-%m-%d')
            fecha_inicio = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen"
        
        # Título
        ws1['A1'] = 'Reporte de Gestión de Laboratorios'
        ws1['A1'].font = Font(size=16, bold=True, color='667eea')
        ws1['A2'] = f'Período: {fecha_inicio} a {fecha_fin}'
        ws1['A3'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        
        # Estadísticas
        ws1['A5'] = 'Métrica'
        ws1['B5'] = 'Valor'
        ws1['A5'].font = Font(bold=True)
        ws1['B5'].font = Font(bold=True)
        ws1['A5'].fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        ws1['B5'].fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        
        query_equipos = "SELECT COUNT(*) as total FROM equipos WHERE estado != 'dado_baja'"
        result = db_manager.ejecutar_query(query_equipos)
        total_equipos = result[0]['total'] if result else 0
        
        query_prestamos = "SELECT COUNT(*) as total FROM prestamos WHERE estado = 'activo'"
        result = db_manager.ejecutar_query(query_prestamos)
        prestamos_activos = result[0]['total'] if result else 0
        
        ws1['A6'] = 'Total Equipos'
        ws1['B6'] = total_equipos
        ws1['A7'] = 'Préstamos Activos'
        ws1['B7'] = prestamos_activos
        
        # Hoja 2: Préstamos
        ws2 = wb.create_sheet("Préstamos")
        headers = ['Código', 'Equipo', 'Solicitante', 'Fecha', 'Estado']
        ws2.append(headers)
        
        # Estilo de encabezados
        for cell in ws2[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        query_prestamos_list = """
            SELECT p.codigo, e.nombre as equipo, 
                   CONCAT(u.nombres, ' ', u.apellidos) as solicitante,
                   DATE_FORMAT(p.fecha, '%d/%m/%Y') as fecha,
                   p.estado
            FROM prestamos p
            JOIN equipos e ON p.id_equipo = e.id
            JOIN usuarios u ON p.id_usuario_solicitante = u.id
            WHERE p.fecha BETWEEN %s AND %s
            ORDER BY p.fecha DESC
            LIMIT 100
        """
        prestamos = db_manager.ejecutar_query(query_prestamos_list, (fecha_inicio, fecha_fin)) or []
        
        for p in prestamos:
            ws2.append([p['codigo'], p['equipo'], p['solicitante'], p['fecha'], p['estado']])
        
        # Ajustar ancho de columnas
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_laboratorios_{fecha_inicio}_{fecha_fin}.xlsx'
        )
        
    except Exception as e:
        print(f"Error generando Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': f'Error: {str(e)}'}, 500

@app.route('/backup', methods=['GET', 'POST'])
def backup():
    """Página de backup del sistema - SOLO ADMINISTRADORES"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar que sea administrador (id_rol = 1)
    user_rol = session.get('user_rol')
    if user_rol != 1:
        flash('Acceso denegado. Solo administradores pueden gestionar backups.', 'error')
        return redirect(url_for('dashboard'))
    
    # Procesar acciones POST
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'create':
            # Crear backup usando la API
            import requests
            try:
                response = requests.post(
                    f"{request.url_root}api/backups/crear",
                    cookies=request.cookies
                )
                data = response.json()
                if data.get('success'):
                    flash(f'Backup creado exitosamente: {data.get("filename")}', 'success')
                else:
                    flash(f'Error creando backup: {data.get("message")}', 'error')
            except Exception as e:
                flash(f'Error creando backup: {str(e)}', 'error')
        
        elif action == 'restore':
            backup_file = request.form.get('backup_file')
            import requests
            try:
                response = requests.post(
                    f"{request.url_root}api/backups/restaurar",
                    json={'filename': backup_file},
                    cookies=request.cookies
                )
                data = response.json()
                if data.get('success'):
                    flash('Base de datos restaurada exitosamente', 'success')
                else:
                    flash(f'Error restaurando backup: {data.get("message")}', 'error')
            except Exception as e:
                flash(f'Error restaurando backup: {str(e)}', 'error')
        
        elif action == 'delete':
            backup_file = request.form.get('backup_file')
            import requests
            try:
                response = requests.post(
                    f"{request.url_root}api/backups/eliminar",
                    json={'filename': backup_file},
                    cookies=request.cookies
                )
                data = response.json()
                if data.get('success'):
                    flash('Backup eliminado exitosamente', 'success')
                else:
                    flash(f'Error eliminando backup: {data.get("message")}', 'error')
            except Exception as e:
                flash(f'Error eliminando backup: {str(e)}', 'error')
        
        return redirect(url_for('backup'))
    
    # Obtener lista de backups
    backups_list = []
    try:
        import requests
        response = requests.get(
            f"{request.url_root}api/backups/listar",
            cookies=request.cookies
        )
        data = response.json()
        if data.get('success'):
            backups_list = data.get('backups', [])
    except Exception as e:
        print(f"Error obteniendo backups: {e}")
    
    return render_template('backup.html', user=get_user_data(), backups=backups_list)

@app.route('/backup/download/<filename>')
def download_backup(filename):
    """Descargar un archivo de backup - SOLO ADMINISTRADORES"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Verificar que sea administrador (id_rol = 1)
    user_rol = session.get('user_rol')
    if user_rol != 1:
        flash('Acceso denegado. Solo administradores pueden descargar backups.', 'error')
        return redirect(url_for('dashboard'))
    
    import os
    from flask import send_file
    
    backup_dir = os.path.join(os.path.dirname(__file__), 'database', 'backups')
    filepath = os.path.join(backup_dir, filename)
    
    if not os.path.exists(filepath):
        flash('Archivo de backup no encontrado', 'error')
        return redirect(url_for('backup'))
    
    try:
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Error descargando backup: {str(e)}', 'error')
        return redirect(url_for('backup'))

@app.route('/configuracion')
def configuracion():
    """Página de configuración del sistema"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if not tiene_permiso('configuracion'):
        flash('No tiene permisos para acceder a configuración.', 'error')
        return redirect(url_for('dashboard'))
    
    # Cargar configuraciones desde la BD
    configuraciones = []
    try:
        query = """
            SELECT clave_config as clave, valor_config as valor, 
                   descripcion, tipo_dato as tipo 
            FROM configuracion_sistema 
            ORDER BY clave_config
        """
        configuraciones = db_manager.ejecutar_query(query) or []
    except Exception as e:
        print(f"Error cargando configuraciones: {e}")
        import traceback
        traceback.print_exc()
    
    return render_template('configuracion.html', user=get_user_data(), configuraciones=configuraciones)

@app.route('/perfil', methods=['GET', 'POST'])
def perfil():
    """Página de perfil del usuario"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    usuario = None
    stats = {'reservas': 0, 'prestamos': 0}
    roles = []
    es_admin = session.get('user_level', 0) >= 4
    
    try:
        query_usuario = """
            SELECT u.id, u.documento, u.nombres, u.apellidos, u.email, u.telefono,
                   u.id_rol, u.fecha_registro, u.ultimo_acceso, u.estado,
                   r.nombre_rol, r.descripcion as rol_descripcion
            FROM usuarios u
            LEFT JOIN roles r ON u.id_rol = r.id
            WHERE u.id = %s
        """
        usuario = db_manager.obtener_uno(query_usuario, (session['user_id'],))
        
        if usuario:
            usuario['nombre_completo'] = f"{usuario['nombres']} {usuario['apellidos']}"
        
        if es_admin:
            query_roles = "SELECT id, nombre_rol FROM roles WHERE estado = 'activo' ORDER BY id"
            roles = db_manager.ejecutar_query(query_roles) or []
        
        try:
            query_stats = "SELECT COUNT(*) as total FROM prestamos WHERE id_usuario_solicitante = %s"
            result = db_manager.obtener_uno(query_stats, (session['user_id'],))
            stats['prestamos'] = result['total'] if result else 0
        except:
            pass
        
        if request.method == 'POST':
            action = request.form.get('action', 'update_profile')
            
            if action == 'update_profile':
                nombres = request.form.get('nombres', '').strip()
                apellidos = request.form.get('apellidos', '').strip()
                email = request.form.get('email', '').strip()
                telefono = request.form.get('telefono', '').strip()
                
                if es_admin:
                    documento = request.form.get('documento', '').strip()
                    nuevo_rol = request.form.get('id_rol')
                    estado = request.form.get('estado', 'activo')
                    
                    query_update = """
                        UPDATE usuarios 
                        SET nombres = %s, apellidos = %s, email = %s, telefono = %s,
                            documento = %s, id_rol = %s, estado = %s
                        WHERE id = %s
                    """
                    db_manager.ejecutar_query(query_update, (
                        nombres, apellidos, email, telefono,
                        documento, nuevo_rol, estado, session['user_id']
                    ))
                else:
                    query_update = "UPDATE usuarios SET nombres = %s, apellidos = %s, email = %s, telefono = %s WHERE id = %s"
                    db_manager.ejecutar_query(query_update, (nombres, apellidos, email, telefono, session['user_id']))
                
                session['user_name'] = f"{nombres} {apellidos}"
                flash('Perfil actualizado correctamente', 'success')
                return redirect(url_for('perfil'))
            
            elif action == 'change_password':
                current_password = request.form.get('current_password', '')
                new_password = request.form.get('new_password', '')
                confirm_password = request.form.get('confirm_password', '')
                
                if new_password != confirm_password:
                    flash('Las contraseñas nuevas no coinciden', 'error')
                    return redirect(url_for('perfil'))
                
                if len(new_password) < 6:
                    flash('La contraseña debe tener al menos 6 caracteres', 'error')
                    return redirect(url_for('perfil'))
                
                query_pass = "SELECT password_hash FROM usuarios WHERE id = %s"
                user_pass = db_manager.obtener_uno(query_pass, (session['user_id'],))
                
                if user_pass and user_pass.get('password_hash'):
                    import bcrypt
                    if not bcrypt.checkpw(current_password.encode('utf-8'), user_pass['password_hash'].encode('utf-8')):
                        flash('La contraseña actual es incorrecta', 'error')
                        return redirect(url_for('perfil'))
                
                import bcrypt
                new_hash = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                
                query_update_pass = "UPDATE usuarios SET password_hash = %s WHERE id = %s"
                db_manager.ejecutar_query(query_update_pass, (new_hash, session['user_id']))
                
                flash('Contraseña actualizada correctamente', 'success')
                return redirect(url_for('perfil'))
        
    except Exception as e:
        print(f"Error en perfil: {e}")
        import traceback
        traceback.print_exc()
    
    if not usuario:
        usuario = {
            'id': session.get('user_id'),
            'documento': session.get('user_id'),
            'nombre_completo': session.get('user_name', 'Usuario'),
            'nombres': session.get('user_name', 'Usuario'),
            'apellidos': '',
            'email': '',
            'telefono': '',
            'id_rol': None,
            'nombre_rol': session.get('user_type', 'Usuario'),
            'estado': 'activo',
            'fecha_registro': None,
            'ultimo_acceso': None
        }
    
    return render_template('perfil.html', 
                           user=get_user_data(), 
                           usuario=usuario, 
                           stats=stats, 
                           roles=roles,
                           es_admin=es_admin)

@app.route('/ayuda')
def ayuda():
    """Página de ayuda del sistema"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('ayuda.html', user=get_user_data())

@app.route('/laboratorio/<int:laboratorio_id>')
def laboratorio_detalle(laboratorio_id):
    """Página de detalle de un laboratorio específico"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    lab = None
    equipos = []
    try:
        query_lab = """
            SELECT l.*, CONCAT(u.nombres, ' ', u.apellidos) as responsable
            FROM laboratorios l
            LEFT JOIN usuarios u ON l.responsable_id = u.id
            WHERE l.id = %s
        """
        lab = db_manager.obtener_uno(query_lab, (laboratorio_id,))
        
        query_equipos = "SELECT * FROM equipos WHERE id_laboratorio = %s ORDER BY nombre"
        equipos = db_manager.ejecutar_query(query_equipos, (laboratorio_id,)) or []
    except Exception as e:
        print(f"Error obteniendo laboratorio: {e}")
        flash('Error al cargar el laboratorio', 'error')
        return redirect(url_for('laboratorios'))
    
    if not lab:
        flash('Laboratorio no encontrado', 'error')
        return redirect(url_for('laboratorios'))
    
    return render_template('laboratorio_detalle.html', user=get_user_data(), lab=lab, equipos=equipos)

@app.route('/admin/solicitudes-nivel')
def gestionar_solicitudes_nivel():
    """Página de gestión de solicitudes de nivel"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    if session.get('user_level', 1) < 6:
        flash('No tiene permisos para acceder a esta página', 'error')
        return redirect(url_for('dashboard'))
    flash('Módulo de solicitudes en desarrollo', 'info')
    return redirect(url_for('dashboard'))

@app.route('/logout')
def logout():
    """Cerrar sesión"""
    session.clear()
    flash('Sesión cerrada exitosamente', 'success')
    return redirect(url_for('login'))

# ========================================
# API ENDPOINTS
# ========================================

@app.route('/api/estadisticas')
def api_estadisticas():
    """API endpoint para obtener estadísticas del dashboard"""
    if 'user_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    # Obtener estadísticas del sistema
    stats = obtener_estadisticas_sistema()
    
    # Mapear estados para la API (mantener compatibilidad con frontend)
    stats_api = stats.copy()
    stats_api['equipos_estado'] = {
        'disponible': stats['equipos_estado'].get('disponible', 0),
        'en_uso': stats['equipos_estado'].get('prestado', 0),
        'mantenimiento': stats['equipos_estado'].get('mantenimiento', 0),
        'fuera_servicio': stats['equipos_estado'].get('reparacion', 0) + stats['equipos_estado'].get('dado_baja', 0)
    }
    
    return jsonify({'estadisticas': stats_api, 'success': True})

# ========================================
# MANEJO DE ERRORES
# ========================================

@app.errorhandler(404)
def not_found(error):
    """Página no encontrada"""
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    """Error interno del servidor"""
    return render_template('500.html'), 500

# ========================================
# CONTEXTO DE PLANTILLAS
# ========================================

@app.context_processor
def inject_user():
    """Inyectar datos del usuario en todas las plantillas"""
    if 'user_id' in session:
        return dict(user=get_user_data())
    return dict(user=None)

# ========================================
# INICIALIZACIÓN
# ========================================

def init_app():
    """Inicializar la aplicación"""
    print("=" * 70)
    print("🔬 SISTEMA DE GESTIÓN INTELIGENTE DE LABORATORIOS (GIL)")
    print("   Centro Minero de Sogamoso - SENA")
    print("=" * 70)
    
    # Verificar conexión a base de datos
    if db_manager.conectar():
        print("✅ Conexión a base de datos exitosa")
    else:
        print("❌ Error conectando a base de datos")
        print("💡 Asegúrese de que MySQL esté corriendo")
        return False
    
    # Crear directorios necesarios
    dirs = ['logs', 'uploads', 'temp']
    for dir_name in dirs:
        dir_path = os.path.join(os.path.dirname(__file__), dir_name)
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
            print(f"📁 Directorio creado: {dir_name}")
    
    print("\n🚀 Servidor iniciando...")
    print("📍 URL: http://localhost:5000")
    print("📍 Login: http://localhost:5000/login")
    print("\n💡 Usuarios de ejemplo:")
    print("   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("   👤 ADMINISTRADORES:")
    print("      • ADMIN001 - Roberto Díaz Silva")
    print("      • TEC_LAB_001 - Téc. Gloria Martínez")
    print("   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("   👨‍🏫 INSTRUCTORES:")
    print("      • INST001 - Carlos Rodríguez Pérez")
    print("      • INST002 - María Elena González")
    print("      • COORD_MIN - Ing. Pedro Sánchez")
    print("   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("   🎓 APRENDICES:")
    print("      • APRE001 - Laura Patricia Ruiz")
    print("      • APRE002 - David Alejandro Castro")
    print("      • MON_001 - John Díaz Pérez")
    print("   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print("   💡 Solo ingrese el ID (sin contraseña)")
    print("=" * 70)
    
    return True

# ========================================
# PUNTO DE ENTRADA
# ========================================

if __name__ == '__main__':
    if init_app():
        # Ejecutar servidor en modo desarrollo
        app.run(
            host='0.0.0.0',
            port=5000,
            debug=True,
            use_reloader=True
        )
    else:
        print("\n❌ No se pudo iniciar el servidor")
        print("Revise la configuración de la base de datos")
        sys.exit(1)